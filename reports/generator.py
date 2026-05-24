"""
Генератор отчётов в форматах: xlsx, csv, json, txt
"""
import os
import json
import csv
import io
from datetime import datetime
from config import REPORTS_DIR

os.makedirs(REPORTS_DIR, exist_ok=True)


def _safe_filename(name: str) -> str:
    import re
    return re.sub(r"[^\w\-]", "_", name)[:40]


async def generate_report(source: dict, items: list[dict], fmt: str) -> tuple[str, int]:
    """
    Генерирует файл отчёта.
    Возвращает (filepath, row_count).
    """
    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    src_name = _safe_filename(source["name"])
    filename = f"{src_name}_{ts}.{fmt}"
    filepath = os.path.join(REPORTS_DIR, filename)

    # Разворачиваем data для плоского вывода
    flat_items = [_flatten(i["data"]) for i in items]

    if fmt == "xlsx":
        _write_xlsx(filepath, flat_items, source)
    elif fmt == "csv":
        _write_csv(filepath, flat_items)
    elif fmt == "json":
        _write_json(filepath, flat_items, source)
    else:
        _write_txt(filepath, flat_items, source)

    return filepath, len(flat_items)


def _flatten(d: dict, prefix: str = "", sep: str = "_") -> dict:
    """Рекурсивно разворачивает вложенные словари."""
    result = {}
    for k, v in d.items():
        key = f"{prefix}{sep}{k}" if prefix else k
        if isinstance(v, dict):
            result.update(_flatten(v, key, sep))
        elif isinstance(v, list):
            result[key] = ", ".join(str(x) for x in v)
        else:
            result[key] = v
    return result


def _write_xlsx(filepath: str, items: list[dict], source: dict):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Данные"

        if not items:
            wb.save(filepath)
            return

        # Заголовки
        headers = list(items[0].keys())
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(color="FFFFFF", bold=True, size=11)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Данные
        alt_fill = PatternFill("solid", fgColor="EBF3FB")
        for row_idx, item in enumerate(items, 2):
            for col, key in enumerate(headers, 1):
                val = item.get(key, "")
                cell = ws.cell(row=row_idx, column=col, value=val)
                if row_idx % 2 == 0:
                    cell.fill = alt_fill
                cell.alignment = Alignment(wrap_text=True)

        # Авто-ширина колонок
        for col in range(1, len(headers) + 1):
            max_len = max(
                len(str(ws.cell(r, col).value or ""))
                for r in range(1, len(items) + 2)
            )
            ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 50)

        # Лист метаданных
        ws_meta = wb.create_sheet("Мета")
        ws_meta["A1"] = "Источник"
        ws_meta["B1"] = source["name"]
        ws_meta["A2"] = "Тип"
        ws_meta["B2"] = source["type"]
        ws_meta["A3"] = "Сгенерирован"
        ws_meta["B3"] = datetime.now().strftime("%d.%m.%Y %H:%M")
        ws_meta["A4"] = "Записей"
        ws_meta["B4"] = len(items)

        wb.save(filepath)
    except ImportError:
        # Fallback если openpyxl не установлен
        _write_csv(filepath.replace(".xlsx", ".csv"), items)


def _write_csv(filepath: str, items: list[dict]):
    if not items:
        open(filepath, "w").close()
        return
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=items[0].keys())
        writer.writeheader()
        writer.writerows(items)


def _write_json(filepath: str, items: list[dict], source: dict):
    payload = {
        "source":    source["name"],
        "type":      source["type"],
        "generated": datetime.now().isoformat(),
        "count":     len(items),
        "data":      items,
    }
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def _write_txt(filepath: str, items: list[dict], source: dict):
    lines = [
        f"=== Отчёт: {source['name']} ===",
        f"Тип: {source['type']}",
        f"Дата: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
        f"Записей: {len(items)}",
        "=" * 40,
        "",
    ]
    for i, item in enumerate(items, 1):
        lines.append(f"[{i}] {item.get('title', '—')}")
        for k, v in item.items():
            if k not in ("title",) and v:
                lines.append(f"  {k}: {v}")
        lines.append("")
    with open(filepath, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
